from typing import Any

import openpyxl as xl  # as gives alias to openpyxl
from pathlib import Path  # P is capital(class)
from openpyxl.chart import BarChart, Reference
# this packacge c
from openpyxl.styles import Alignment

# let's define an automated script works with every xsl @line 14

directory = Path("transactions.xlsx")
# instead of typing openpyxl. we will use xl.

# wb = xl.load_workbook(Path=\"C:\Users\Joseph\PycharmProjects\transactions.xlsx))
filename = str(input("Enter the filename of the excel"))


def process_workbook(filename):
    wb = xl.load_workbook(filename)  # this will return a workbook obj
    sheet = wb['Sheet1'] # to access the sheet1 specify the name:
    sheet.title = "Transactions"
    # first_cell=sheet['a1']
    # =sheet.cell(row=1,column=1)
    # first_cell=sheet.cell(1,1)   #addressses A1

    # print(first_cell.value)   # output:transaction_id

    # print(first_cell)   #output:<Cell 'Transactions'.A1>
    # print(sheet.max_row)
    # print(sheet.max_column)

    # in range: adding 1 will start
    # from 1 to # include last row  since RANGE method does not include last param
    # for row in range(1,sheet.max_row+1):
    #     print(row )
    # we want to ignore first row since it's headlines
    for row in range(2, sheet.max_row + 1):
        price = sheet.cell(row, 3)
        # print(cell.value)
        new_price = price.value * 0.9
        new_price_cell = sheet.cell(row, 4)  # new cell id which is new column D(4th)
        new_price_cell.value = new_price
        print(new_price)
    wb.save(filename)
    # now we need to import charts  and create charts for this excel @LINE3

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    wb.save(filename)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')
    wb.save(filename)  # created with a default chart


# "to apply process_workbook function to all sheets:"
process_workbook(filename)
